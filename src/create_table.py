from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def print_rows(ws):
    row_strings = ''

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            max_col=ws.max_column):
        ws.iter_rows()
        for cell in row:
            row_strings += "{:<6}".format(str(cell.value) + " ")
        row_strings += "\n"
    print(row_strings)


if __name__ == '__main__':
    filename = 'create_table.xlsx'
    wb = Workbook()
    ws1 = wb.worksheets[0]

    sales_data = [['Medellin', 1200, 1500],
                  ['Bogota', 900, 1000],
                  ['Barranquilla', 3000, 3200],
                  ['Cali', 500, 300]]

    ws1.append(["SALES", "2023", "2024"])
    for row in sales_data:
        ws1.append(row)

    print_rows(ws1)

    ##Create Excel Table
    sales_table = Table(displayName='SalesTable', ref='A1:C5')

    style = TableStyleInfo(name='TableStyleMedium8', showRowStripes=True,
                           showColumnStripes=True)

    sales_table.tableStyleInfo = style
    ws1.add_table(sales_table)

    wb.save(filename)
