import openpyxl

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

ws.cell(row=3, column=3).value = 'AmountUSD'
ws['C4'] = 32.05

ws['A2'] = 'Hello'
ws['B2'] = 'World'

wb.save('add_hello_world.xlsx')