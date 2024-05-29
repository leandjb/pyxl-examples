from openpyxl import Workbook

wb = Workbook()
wb.remove(wb["Sheet"])
ws1 = wb.create_sheet("xheet_1", 0)
ws2 = wb.create_sheet("xheet_2", 1)

for sheet in wb:
    print(sheet.title)

wb.save("example2.xlsx")