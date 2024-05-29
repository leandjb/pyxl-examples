import openpyxl


def set_values(ws):
    ws.delete_cols(1, 100)
    counter = 1

    for row in ws.iter_rows(min_row=1, max_row=10, max_col=10):
        for cell in row:
            cell.value = counter
            counter += 1


def print_rows(ws):
    row_strings = ''

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            max_col=ws.max_column):
        ws.iter_rows()
        for cell in row:
            row_strings += "{:<3}".format(str(cell.value) + " ")
        row_strings += "\n"
    print(row_strings)


if __name__ == '__main__':
    filename = 'add_data_rows.xlsx'
    wb = openpyxl.Workbook()
    ws1 = wb.worksheets[0]

    set_values(ws1)
    print_rows(ws1)

    ws1.insert_rows(0)
    ws1.insert_rows(5)
    ws1.delete_rows(5)

    wb.save('add_data_rows.xlsx')
